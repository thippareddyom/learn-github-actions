from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import requests

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit("openpyxl is required; install with `pip install openpyxl`") from exc

URL_TEMPLATE = "https://www.ssga.com/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{ticker}.xlsx"
REPO_ROOT = Path(__file__).resolve().parents[1]
SERVICE_ROOT = REPO_ROOT / "ark_list_web_service"
HOLDINGS_DIR = SERVICE_ROOT / "data" / "holdings"
HOLDINGS_URLS = SERVICE_ROOT / "configs" / "holdings_urls.json"


def download_xlsx(url: str, dest: Path) -> Path:
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError(f"Download failed for {url}: {exc}") from exc
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(resp.content)
    return dest


def _is_header(row: List[str]) -> bool:
    lowered = [str(c or "").strip().lower() for c in row]
    has_ticker = any(c in ("ticker", "symbol") for c in lowered)
    has_weight = any("weight" in c for c in lowered)
    return has_ticker and has_weight


def parse_holdings(xlsx_path: Path) -> List[Dict[str, object]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = [[cell for cell in r] for r in ws.iter_rows(values_only=True)]
    header_idx = next((i for i, r in enumerate(rows) if _is_header([str(c or "") for c in r])), None)
    if header_idx is None:
        raise RuntimeError("Could not locate header row with Ticker/Symbol and Weight columns")

    headers = [str(c or "").strip() for c in rows[header_idx]]
    records: List[Dict[str, object]] = []
    for r in rows[header_idx + 1 :]:
        if not any(r):
            continue
        values = list(r)
        if len(values) < len(headers):
            values += [None] * (len(headers) - len(values))
        record = {headers[i]: values[i] for i in range(len(headers))}
        records.append(record)
    return records


def save_json(path: Path, rows: List[Dict[str, object]], source: str, symbol: str) -> None:
    payload = {
        "symbol": symbol,
        "source": source,
        "downloaded_at": datetime.utcnow().isoformat() + "Z",
        "count": len(rows),
        "rows": rows,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2))


def resolve_url(ticker: str) -> str:
    """Prefer configured URL in holdings_urls.json, otherwise build from template."""
    ticker_upper = ticker.upper()
    if HOLDINGS_URLS.exists():
        try:
            data = json.loads(HOLDINGS_URLS.read_text())
            url = data.get(ticker_upper)
            if isinstance(url, str) and url.strip():
                return url.strip()
        except Exception:
            pass
    return URL_TEMPLATE.format(ticker=ticker.lower())


def main(ticker: str = "SPY") -> int:
    ticker = ticker.upper()
    url = resolve_url(ticker)
    xlsx_path = download_xlsx(url, REPO_ROOT / f"{ticker}_holdings.xlsx")
    holdings = parse_holdings(xlsx_path)
    json_path = HOLDINGS_DIR / f"{ticker}-holdings.json"
    save_json(json_path, holdings, url, ticker)
    print(f"Saved {len(holdings)} rows to {json_path}")
    return 0


if __name__ == "__main__":
    import sys

    sym = sys.argv[1] if len(sys.argv) > 1 else "SPY"
    raise SystemExit(main(sym))
